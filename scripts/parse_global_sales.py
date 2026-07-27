#!/usr/bin/env python3
"""
Parse the daily company-wide sales export ("data penjualan global ...xlsx") into data.json for
the Kuta-Seminyak dashboard.

The company export is a flat transaction-level sheet covering every region/rep/customer in the
business. This script filters it down to Lani's now-unstaffed accounts — the "Sales" column carries
the literal placeholder "Vacant Hotel Kuta Seminyak" / "Vacant Hotel Seminyak Kuta" for that
territory — aggregates by customer -> product -> month, and writes the same data.json shape
build_dashboard.py already consumes.

Usage:
    python3 parse_global_sales.py [path/to/data_penjualan_global.xlsx] [path/to/output.json]

Defaults to the newest-looking "*.xlsx" next to this project (by the "jan-<mon>" suffix in the
filename) and writes "data.json" in the project root.
"""
import sys
import re
import json
from pathlib import Path
from datetime import datetime
from collections import defaultdict

import openpyxl

MONTHLY_TARGET = 1_000_000_000  # 1B IDR/month, the long-term growth target

REQUIRED_COLUMNS = ['Nama Cust', 'Nama Brg', 'Kategori Barang', 'Grand Total', 'Sales', 'bulan', 'nama bulan', 'tahun']

_MONTH_ABBR = {'jan': 1, 'feb': 2, 'mar': 3, 'apr': 4, 'may': 5, 'jun': 6,
               'jul': 7, 'aug': 8, 'sep': 9, 'oct': 10, 'nov': 11, 'dec': 12}


def find_default_xlsx(project_root: Path) -> Path:
    candidates = list(project_root.glob("*.xlsx"))
    if not candidates:
        sys.exit(f"No .xlsx file found in {project_root}")
    if len(candidates) == 1:
        return candidates[0]

    def key(p):
        m = re.search(r'jan[-\s]?(\w{3})', p.name.lower())
        return _MONTH_ABBR.get(m.group(1), 0) if m else 0

    return max(candidates, key=key)


def load_monthly_targets(project_root: Path) -> tuple[int, dict]:
    """Reads monthly_targets.json (default long-term target + a per-month ramp override map
    keyed 'YYYY-MM'). Falls back to the flat MONTHLY_TARGET if the file is missing."""
    path = project_root / "monthly_targets.json"
    if not path.exists():
        return MONTHLY_TARGET, {}
    cfg = json.loads(path.read_text())
    return int(cfg.get("default_target", MONTHLY_TARGET)), cfg.get("targets", {})


def is_vacant_kuta_seminyak(sales_value) -> bool:
    s = str(sales_value or '').strip().lower()
    return 'vacant' in s and 'kuta' in s and 'seminyak' in s


def normalize_category(raw) -> str:
    s = str(raw or '').strip()
    return s.upper().title() if s else 'Other'


def load_rows(xlsx_path: Path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter)
    col = {name: i for i, name in enumerate(header) if name is not None}
    missing = [c for c in REQUIRED_COLUMNS if c not in col]
    if missing:
        sys.exit(f"Missing expected column(s) in {xlsx_path.name}: {missing}")
    return col, list(rows_iter)


def detect_month_order_and_partial(col, rows):
    """Use the FULL (unfiltered) dataset for a statistically meaningful partial-month signal —
    our filtered subset is too small (~1-2k rows) to tell a partial month from a quiet one."""
    month_names = {}
    counts_all = defaultdict(int)
    for r in rows:
        b, y, nm = r[col['bulan']], r[col['tahun']], r[col['nama bulan']]
        if b is None:
            continue
        key = (int(y or 0), int(b))
        counts_all[key] += 1
        month_names.setdefault(key, str(nm).strip() if nm else str(b))

    months_sorted = sorted(counts_all)
    if not months_sorted:
        sys.exit("No month data found in the export.")

    latest = months_sorted[-1]
    prior = months_sorted[:-1]
    avg_prior = (sum(counts_all[m] for m in prior) / len(prior)) if prior else counts_all[latest]
    partial = latest if counts_all[latest] < avg_prior * 0.75 else None

    return months_sorted, month_names, partial


def main():
    project_root = Path(__file__).resolve().parent.parent
    xlsx_path = Path(sys.argv[1]) if len(sys.argv) > 1 else find_default_xlsx(project_root)
    out_path = Path(sys.argv[2]) if len(sys.argv) > 2 else project_root / "data.json"

    col, rows = load_rows(xlsx_path)
    month_keys, month_names, partial_key = detect_month_order_and_partial(col, rows)
    months = [month_names[k] + ('*' if k == partial_key else '') for k in month_keys]
    key_index = {k: i for i, k in enumerate(month_keys)}
    n_months = len(months)
    trend_len = n_months - 1 if partial_key is not None else n_months

    default_target, target_overrides = load_monthly_targets(project_root)
    monthly_targets = [
        int(target_overrides.get(f"{y:04d}-{m:02d}", default_target)) for y, m in month_keys
    ]

    vacant_rows = [r for r in rows if is_vacant_kuta_seminyak(r[col['Sales']])]

    customers = {}
    for r in vacant_rows:
        b, y = r[col['bulan']], r[col['tahun']]
        if b is None:
            continue
        mi = key_index.get((int(y or 0), int(b)))
        if mi is None:
            continue

        cname = str(r[col['Nama Cust']] or '').strip()
        pname = str(r[col['Nama Brg']] or '').strip()
        if not cname or not pname:
            continue
        amount = float(r[col['Grand Total']] or 0)
        category = normalize_category(r[col['Kategori Barang']])

        c = customers.setdefault(cname, {"name": cname, "monthly": [0.0] * n_months, "products": {}})
        c["monthly"][mi] += amount

        p = c["products"].setdefault(pname, {"name": pname, "category": category, "monthly": [0.0] * n_months})
        p["monthly"][mi] += amount

    for c in customers.values():
        c["products"] = list(c["products"].values())
        for p in c["products"]:
            p["total"] = sum(p["monthly"])
        c["total"] = sum(c["monthly"])

    data = build_data(months, list(customers.values()), trend_len, partial_key is not None,
                       default_target, monthly_targets)
    out_path.write_text(json.dumps(data, indent=2, ensure_ascii=False))

    print(f"Source: {xlsx_path.name}")
    print(f"Filtered {len(vacant_rows):,} of {len(rows):,} rows (Vacant Hotel Kuta Seminyak / Seminyak Kuta)")
    print(f"Parsed {data['customer_count']} customers, {data['product_count']} products, "
          f"{n_months} months ({months[0]}-{months[-1]})")
    if data['partial_month']:
        print(f"Partial month detected: {data['partial_month']} (excluded from run-rate/trend)")
    print(f"Grand total: {data['grand_total']:,.0f}")
    print(f"Run rate (last {min(3, trend_len)} complete mo avg): {data['run_rate']:,.0f} "
          f"({data['target_progress_pct']:.1f}% of {default_target:,.0f} long-term target)")
    print(f"This month ({months[-1]}): {data['current_month_actual']:,.0f} of "
          f"{data['current_month_target']:,.0f} target ({data['current_month_progress_pct']:.1f}%)")
    print(f"At-risk customers: {len(data['at_risk'])}, Upsell targets: {len(data['upsell_targets'])}")
    print(f"Wrote {out_path}")


def trend_status(monthly, window=3):
    n = len(monthly)
    if n == 0:
        return "flat"
    if monthly[-1] == 0 and sum(monthly[:-1]) > 0:
        return "churned"
    w = min(window, n)
    recent = monthly[-w:]
    prior = monthly[:-w]
    recent_avg = sum(recent) / len(recent) if recent else 0
    if not prior:
        return "growing" if recent_avg > 0 else "flat"
    prior_avg = sum(prior) / len(prior)
    if prior_avg == 0:
        return "new" if recent_avg > 0 else "flat"
    if recent_avg > prior_avg * 1.15:
        return "growing"
    if recent_avg < prior_avg * 0.85:
        return "declining"
    return "flat"


def build_data(months, customers, trend_len, has_partial, default_target, monthly_targets):
    grand_total = sum(c["total"] for c in customers)

    monthly_totals = [0.0] * len(months)
    for c in customers:
        for i, v in enumerate(c["monthly"]):
            monthly_totals[i] += v

    run_rate_window = min(3, trend_len)
    run_rate = sum(monthly_totals[:trend_len][-run_rate_window:]) / run_rate_window if run_rate_window else 0

    ranked = sorted(customers, key=lambda c: -c["total"])
    for i, c in enumerate(ranked):
        c["rank"] = i + 1
        c["pct_of_total"] = (c["total"] / grand_total * 100) if grand_total else 0
        c["product_count"] = len(c["products"])
        trend_monthly = c["monthly"][:trend_len] if has_partial else c["monthly"]
        c["status"] = trend_status(trend_monthly)
        last_idx = max((i for i, v in enumerate(c["monthly"]) if v > 0), default=None)
        c["last_active_month"] = months[last_idx] if last_idx is not None else None

    top5_total = sum(c["total"] for c in ranked[:5])
    top10_total = sum(c["total"] for c in ranked[:10])

    at_risk = sorted([c for c in ranked if c["status"] == "churned"], key=lambda c: -c["total"])
    upsell_targets = sorted(
        [c for c in ranked if c["product_count"] <= 2 and c["status"] != "churned"],
        key=lambda c: -c["total"],
    )

    product_agg = {}
    for c in ranked:
        for p in c["products"]:
            entry = product_agg.setdefault(p["name"], {
                "name": p["name"], "category": p["category"], "total": 0, "customers": set(),
            })
            entry["total"] += p["total"]
            entry["customers"].add(c["name"])
    products = sorted(product_agg.values(), key=lambda p: -p["total"])
    for p in products:
        p["customer_count"] = len(p["customers"])
        del p["customers"]

    category_totals = {}
    for p in products:
        category_totals[p["category"]] = category_totals.get(p["category"], 0) + p["total"]
    categories = sorted(
        [{"category": k, "total": v} for k, v in category_totals.items()],
        key=lambda x: -x["total"],
    )

    return {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "months": months,
        "partial_month": months[-1] if has_partial else None,
        "monthly_totals": monthly_totals,
        "grand_total": grand_total,
        "run_rate": run_rate,
        "monthly_target": default_target,
        "target_progress_pct": (run_rate / default_target * 100) if default_target else 0,
        "monthly_targets": monthly_targets,
        "current_month_target": monthly_targets[-1],
        "current_month_actual": monthly_totals[-1],
        "current_month_progress_pct": (
            (monthly_totals[-1] / monthly_targets[-1] * 100) if monthly_targets[-1] else 0
        ),
        "customer_count": len(ranked),
        "product_count": len(products),
        "top5_pct_of_total": (top5_total / grand_total * 100) if grand_total else 0,
        "top10_pct_of_total": (top10_total / grand_total * 100) if grand_total else 0,
        "customers": ranked,
        "at_risk": at_risk,
        "upsell_targets": upsell_targets,
        "products": products,
        "categories": categories,
    }


if __name__ == "__main__":
    main()
