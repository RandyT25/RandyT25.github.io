#!/usr/bin/env python3
"""
Parse a Lani-Customer-style Excel pivot export into clean data.json for the sales dashboard.

Usage:
    python3 parse_sales.py [path/to/export.xlsx] [path/to/output.json]

Defaults to "data Lani Customer.xlsx" in the project root and writes "data.json" next to it.

The source file is a single-sheet Excel pivot table export where column A alternates between:
  - customer rows   (indent level 0, e.g. "PADMA BALI RESORT")
  - product rows    (indent level 1, nested under the customer directly above)
Columns after the label are one per month, followed by a "Grand Total" column, ending in a
"Grand Total" row. Month columns are read from the header row so this keeps working if next
month's export adds a new column (e.g. August).
"""
import sys
import json
from pathlib import Path
from datetime import datetime

import openpyxl

MONTHLY_TARGET = 1_000_000_000  # 1B IDR/month, the long-term growth target

CATEGORY_KEYWORDS = [
    ("Seafood", ["crab", "shrimp", "prawn", "salmon", "tuna", "lobster", "squid", "fish"]),
    ("Dairy", ["mozzarella", "butter", "cream", "cheese", "milk", "yogurt", "yoghurt"]),
    ("Beverage & Tea", ["tea", "matcha", "coffee", "juice", "soda", "water", "glass still"]),
    ("Flour & Dry Goods", ["tepung", "flour", "pasta", "linguine", "spaghetti", "penne"]),
    ("Meat & Protein", [
        "tenderloin", "topside", "oxtail", "duck", "beef", "chicken", "sirloin",
        "striploin", "ribeye", "wagyu", "sausage", "bacon", "ham", "lamb", "pork",
    ]),
    ("Bakery & Pastry", ["bread", "croissant", "cake", "hash brown", "wedges"]),
]


def categorize(product_name: str) -> str:
    lower = product_name.lower()
    for category, keywords in CATEGORY_KEYWORDS:
        if any(kw in lower for kw in keywords):
            return category
    return "Other"


def parse_workbook(xlsx_path: Path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    # Find the header row (contains "Row Labels" in column A) and the month columns
    # between it and the "Grand Total" column.
    header_row = None
    for r in range(1, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == "Row Labels":
            header_row = r
            break
    if header_row is None:
        raise ValueError('Could not find header row ("Row Labels") in sheet')

    months = []
    grand_total_col = None
    col = 2
    while True:
        val = ws.cell(row=header_row, column=col).value
        if val is None:
            break
        if str(val).strip().lower() == "grand total":
            grand_total_col = col
            break
        months.append(str(val).strip())
        col += 1
    if grand_total_col is None:
        raise ValueError('Could not find "Grand Total" column in header row')

    customers = []
    cur_customer = None
    for r in range(header_row + 1, ws.max_row + 1):
        cell = ws.cell(row=r, column=1)
        name = cell.value
        if name is None:
            continue
        if str(name).strip().lower() == "grand total":
            break  # final grand total row
        indent = cell.alignment.indent or 0.0
        monthly = [ws.cell(row=r, column=2 + i).value or 0 for i in range(len(months))]
        total = ws.cell(row=r, column=grand_total_col).value or 0
        if indent == 0.0:
            cur_customer = {"name": str(name).strip(), "monthly": monthly, "total": total, "products": []}
            customers.append(cur_customer)
        else:
            if cur_customer is None:
                continue
            cur_customer["products"].append({
                "name": str(name).strip(),
                "category": categorize(str(name)),
                "monthly": monthly,
                "total": total,
            })

    return months, customers


def trend_status(monthly, window=3):
    n = len(monthly)
    if n == 0:
        return "flat"
    if monthly[-1] == 0 and sum(monthly[:-1]) > 0:
        return "churned"
    w = min(window, n)
    recent = monthly[-w:]
    prior = monthly[:-w]
    recent_avg = sum(recent) / len(recent) if recent else 0
    if not prior:
        return "growing" if recent_avg > 0 else "flat"
    prior_avg = sum(prior) / len(prior)
    if prior_avg == 0:
        return "new" if recent_avg > 0 else "flat"
    if recent_avg > prior_avg * 1.15:
        return "growing"
    if recent_avg < prior_avg * 0.85:
        return "declining"
    return "flat"


def build_data(months, customers):
    grand_total = sum(c["total"] for c in customers)

    monthly_totals = [0] * len(months)
    for c in customers:
        for i, v in enumerate(c["monthly"]):
            monthly_totals[i] += v

    run_rate_window = min(3, len(months))
    run_rate = sum(monthly_totals[-run_rate_window:]) / run_rate_window if run_rate_window else 0

    ranked = sorted(customers, key=lambda c: -c["total"])
    for i, c in enumerate(ranked):
        c["rank"] = i + 1
        c["pct_of_total"] = (c["total"] / grand_total * 100) if grand_total else 0
        c["product_count"] = len(c["products"])
        c["status"] = trend_status(c["monthly"])
        # last active month = last index with nonzero revenue
        last_idx = max((i for i, v in enumerate(c["monthly"]) if v > 0), default=None)
        c["last_active_month"] = months[last_idx] if last_idx is not None else None

    top5_total = sum(c["total"] for c in ranked[:5])
    top10_total = sum(c["total"] for c in ranked[:10])

    at_risk = sorted(
        [c for c in ranked if c["status"] == "churned"],
        key=lambda c: -c["total"],
    )
    upsell_targets = sorted(
        [c for c in ranked if c["product_count"] <= 2 and c["status"] != "churned"],
        key=lambda c: -c["total"],
    )

    # Product performance across all customers
    product_agg = {}
    for c in ranked:
        for p in c["products"]:
            entry = product_agg.setdefault(p["name"], {
                "name": p["name"], "category": p["category"], "total": 0, "customers": set(),
            })
            entry["total"] += p["total"]
            entry["customers"].add(c["name"])
    products = sorted(product_agg.values(), key=lambda p: -p["total"])
    for p in products:
        p["customer_count"] = len(p["customers"])
        del p["customers"]

    category_totals = {}
    for p in products:
        category_totals[p["category"]] = category_totals.get(p["category"], 0) + p["total"]
    categories = sorted(
        [{"category": k, "total": v} for k, v in category_totals.items()],
        key=lambda x: -x["total"],
    )

    return {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "months": months,
        "monthly_totals": monthly_totals,
        "grand_total": grand_total,
        "run_rate": run_rate,
        "monthly_target": MONTHLY_TARGET,
        "target_progress_pct": (run_rate / MONTHLY_TARGET * 100) if MONTHLY_TARGET else 0,
        "monthly_targets": [MONTHLY_TARGET] * len(months),
        "current_month_target": MONTHLY_TARGET,
        "current_month_actual": monthly_totals[-1] if monthly_totals else 0,
        "current_month_progress_pct": (
            (monthly_totals[-1] / MONTHLY_TARGET * 100) if monthly_totals and MONTHLY_TARGET else 0
        ),
        "customer_count": len(ranked),
        "product_count": len(products),
        "top5_pct_of_total": (top5_total / grand_total * 100) if grand_total else 0,
        "top10_pct_of_total": (top10_total / grand_total * 100) if grand_total else 0,
        "customers": ranked,
        "at_risk": at_risk,
        "upsell_targets": upsell_targets,
        "products": products,
        "categories": categories,
    }


def main():
    project_root = Path(__file__).resolve().parent.parent
    xlsx_path = Path(sys.argv[1]) if len(sys.argv) > 1 else project_root / "data Lani Customer.xlsx"
    out_path = Path(sys.argv[2]) if len(sys.argv) > 2 else project_root / "data.json"

    months, customers = parse_workbook(xlsx_path)
    data = build_data(months, customers)

    out_path.write_text(json.dumps(data, indent=2, ensure_ascii=False))

    print(f"Parsed {data['customer_count']} customers, {data['product_count']} products, "
          f"{len(months)} months ({months[0]}-{months[-1]})")
    print(f"Grand total: {data['grand_total']:,.0f}")
    print(f"Run rate (last {min(3, len(months))} mo avg): {data['run_rate']:,.0f} "
          f"({data['target_progress_pct']:.1f}% of {MONTHLY_TARGET:,.0f} target)")
    print(f"At-risk customers: {len(data['at_risk'])}, Upsell targets: {len(data['upsell_targets'])}")
    print(f"Wrote {out_path}")


if __name__ == "__main__":
    main()
